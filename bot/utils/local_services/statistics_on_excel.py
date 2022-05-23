import datetime
import pytz
import logging

from openpyxl import load_workbook
from openpyxl.workbook import Workbook

from loader import db


def get_timezone():
    return pytz.timezone('Asia/Tashkent')


def get_current_time():
    now = datetime.datetime.now(get_timezone())
    return now.strftime('%Y%m%d_%H%M')


async def get_excel_file_with_data(chat_id: str, basin_id: str, start_date: datetime = None, end_date: datetime = None):
    file_name = f'{chat_id}_{get_current_time()}.xlsx'

    try:
        wb = load_workbook(file_name)
    except FileNotFoundError as err:
        logging.info(f'{err}. Creating new file.')
        wb = Workbook()

    wb.remove_sheet(wb.active)

    try:
        ws = wb.get_sheet_by_name(basin_id)
    except KeyError as err:
        logging.info(f'{err}. Creating new sheet')
        ws = wb.create_sheet(basin_id)

    ws['A1'] = 'Sana'
    ws['B1'] = 'Vaqt'
    ws['C1'] = 'Metr kub/soat'

    if start_date is not None and end_date is not None:
        messages = await db.get_basin_messages_between_dates(basin_id, start_date, end_date)
    elif start_date is not None:
        messages = await db.get_basin_messages_after_date(basin_id, start_date)
    elif end_date is not None:
        messages = await db.get_basin_messages_before_date(basin_id, end_date)
    else:
        messages = await db.get_basin_messages(basin_id)

    counter = 2
    for message in messages:
        dt = message[10] + datetime.timedelta(hours=5)
        date = dt.strftime("%Y.%m.%d")
        time = dt.strftime("%H:%M")
        ws[f'A{counter}'] = date
        ws[f'B{counter}'] = time
        ws[f'C{counter}'] = message[4]
        counter += 1

    wb.save(file_name)
    return file_name
